#!/usr/bin/python3
# -*- coding: utf-8 -*-

import openpyxl as xl
from openpyxl import load_workbook
from docx import Document
import pandas as pd
from tkinter import *
import subprocess

import Check_List





root = Tk()
root.title("Меню")
name = StringVar()
name2 = StringVar()
name_label2 = Label(text="Введите путь и название файла (Пример - C:/Users/Данные/АСУ.xlsx) :")
name_label2.grid(row=0, column=0, sticky="w")
name_entry2 = Entry(textvariable=name2,width=80)
name_entry2.grid(row=2, column=1, padx=5, pady=10)


name_label = Label(text="Введите путь до расположения паспортов (Пример C:/Users/ИД_РЦ ) :")
name_label.grid(row=2, column=0, sticky="w")
name_entry = Entry(textvariable=name,width=80)
name_entry.grid(row=0, column=1, padx=5, pady=5)
message_button = Button(text="Click Me")
message_button.grid(row=3,column=1, padx=5, pady=5, sticky="e")

root.mainloop()

asutp = name.get()
doc_path = name2.get()

filename = asutp
book = load_workbook(filename)
ws = book.worksheets[1]
for cell in ws["A"]:
    if cell.value is None:
        cl= cell.row
        print(cl)
        break

for d in range(1,cl):
    wb = xl.load_workbook(filename)
    sheets = wb.sheetnames
    dnew= str(d)
    jo_A=''.join(['A',dnew])
    jo_B = ''.join(['B', dnew])
    doc_val = wb[sheets[1]][jo_A].value
    name_Asu = wb[sheets[1]][jo_B].value

    first_cell_string = 'Производитель'
    name_doc= doc_path + '/' + doc_val
    document = Document(name_doc)
    # create a list of all of the table object with text of the
    # first cell equal to `first_cell_string`
    tables = [t for t in document.tables
              if t.cell(0, 1).text.strip() == first_cell_string]

    # in the case that more than one table is found
    for table in tables:
        data = []
        for i, row in enumerate(table.rows):
            text = (cell.text for cell in row.cells)
            if i == 0:
                keys = tuple(text)
                continue

            row_data = dict(zip(keys, text))
            data.append(row_data)
        df = pd.DataFrame(data)
        with pd.ExcelWriter("output.xlsx", engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, startrow=0, startcol=0)
    Check_List.check_list(name_Asu)

subprocess.call("NNSTDASU_шаблон.xlsx", shell=True)