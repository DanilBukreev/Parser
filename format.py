import test4_del_sheets
from openpyxl import load_workbook


def format():
    wb = load_workbook('output.xlsx')
    sheets = wb.sheetnames
    if len(sheets) == 4:
        ws1 = wb.worksheets[1]
        ws2 = wb.worksheets[2]
        ws3 = wb.worksheets[3]
        ws1.delete_cols(1, 2)
        ws1.delete_cols(2, 3)
        ws1.delete_rows(1)

        ws2.delete_cols(1, 2)
        ws2.delete_cols(3, 5)
        ws2.delete_rows(1)

        ws3.delete_cols(1, 2)
        ws3.delete_cols(3, 5)
        ws3.delete_rows(1)

        wb.save('output.xlsx')

    elif len(sheets) == 3:
        ws1 = wb.worksheets[1]
        ws2 = wb.worksheets[2]
        ws1.delete_cols(1, 2)
        ws1.delete_cols(2, 3)
        ws1.delete_rows(1)

        ws2.delete_cols(1, 2)
        ws2.delete_cols(3, 5)
        ws2.delete_rows(1)

        wb.save('output.xlsx')

    elif len(sheets) == 2:
        ws1 = wb.worksheets[1]
        ws1.delete_cols(1, 2)
        ws1.delete_cols(2, 3)
        ws1.delete_rows(1)

        wb.save('output.xlsx')


