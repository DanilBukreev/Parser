#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
'''
thin_border = Border(top=Side(style='medium'))

wb = load_workbook('output.xlsx')
ws = wb.worksheets[0]
# property cell.border should be used instead of cell.style.border
cel_row=5
for i in range(1,5):
    ws.cell(row=cel_row, column=i).border = thin_border
    wb.save('output.xlsx')
'''
book = load_workbook("NNSTDASU_шаблон.xlsx")
ws = book.worksheets[1]
for cell in ws["C"]:
    if cell.value is None:
        print(cell.row)
        break
# cell format underline last full row
thin_border = Border(top=Side(style='medium'))
for i in range(1, 6):
    ws.cell(row=cell.row, column=i).border = thin_border
    book.save('NNSTDASU_шаблон.xlsx')