#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import openpyxl as xl

def copy_data():
    out = load_workbook("output.xlsx")
    sheets = out.sheetnames
    for k in range(1,len(sheets)):
        book = load_workbook("NNSTDASU_шаблон.xlsx")
        ws = book.worksheets[k]
        for cell in ws["C"]:
            if cell.value is None:
                print(cell.row)
                break

        # opening the source excel file
        filename = "output.xlsx"
        wb1 = xl.load_workbook(filename)
        ws1 = wb1.worksheets[k]

        # opening the destination excel file
        filename1 = "NNSTDASU_шаблон.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.worksheets[k]

        # calculate total number of rows and
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column

        # copying the cell values from source
        # excel file to destination excel file

        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = ws1.cell(row=i, column=j)

                # writing the read value to destination excel file
                ws2.cell(row=i+cell.row-1, column=j+2).value = c.value
        print("C",cell.row)
        # saving the destination excel file
        wb2.save(str(filename1))







