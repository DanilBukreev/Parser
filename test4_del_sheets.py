from openpyxl import load_workbook

def del_sheet():
    wb = load_workbook('output.xlsx')
    sheets = wb.sheetnames
    for k in range (1,len(sheets)):
        wb.remove(wb[sheets[k]])
        wb.save('output.xlsx')


def del_sheet_k(k):
    wb = load_workbook('output.xlsx')
    sheets = wb.sheetnames
    wb.remove(wb[sheets[k]])
    wb.save('output.xlsx')

