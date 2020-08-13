#!/usr/bin/python3
# -*- coding: utf-8 -*-


import openpyxl as xl
from openpyxl import load_workbook

def insert_up(doc):
    book = load_workbook("NNSTDASU_шаблон.xlsx")
    ws = book.worksheets[1]
    for cell in ws["C"]:
        if cell.value is None:
            cl=cell.row
            print(cl)
            break
    filename = 'АСУ.xlsx'
    wb = xl.load_workbook(filename)
    sheets = wb.sheetnames
    dnew= str(doc)
    join_d=''.join(['C',dnew])
    name_Asu = wb[sheets[1]][join_d].value
    print(name_Asu)

    cl_Asu=str(cell.row)
    join_cl=''.join(['B',cl_Asu])
    wb2 = xl.load_workbook('NNSTDASU_шаблон.xlsx')
    sheet= wb2.worksheets[1]
    sheet[join_cl] = name_Asu
    wb2.save('NNSTDASU_шаблон.xlsx')


def insert_mid(doc):
    book = load_workbook("NNSTDASU_шаблон.xlsx")
    ws = book.worksheets[2]
    for cell in ws["C"]:
        if cell.value is None:
            cl=cell.row
            print(cl)
            break
    filename = 'АСУ.xlsx'
    wb = xl.load_workbook(filename)
    sheets = wb.sheetnames
    dnew= str(doc)
    join_d=''.join(['C',dnew])
    name_Asu = wb[sheets[1]][join_d].value
    print(name_Asu)

    cl_Asu=str(cell.row)
    join_cl=''.join(['B',cl_Asu])
    wb2 = xl.load_workbook('NNSTDASU_шаблон.xlsx')
    sheet= wb2.worksheets[2]
    sheet[join_cl] = name_Asu
    wb2.save('NNSTDASU_шаблон.xlsx')


def insert_set(doc):
    book = load_workbook("NNSTDASU_шаблон.xlsx")
    ws = book.worksheets[3]
    for cell in ws["C"]:
        if cell.value is None:
            cl=cell.row
            print(cl)
            break
    filename_ASU = 'АСУ.xlsx'
    wb_up = xl.load_workbook(filename_ASU)
    sheets_up = wb_up.sheetnames
    dnew = str(doc)
    join_d = ''.join(['C', dnew])
    name_Asu = wb_up[sheets_up[1]][join_d].value
    print(name_Asu)

    cl_Asu = str(cell.row)
    join_cl = ''.join(['B', cl_Asu])
    wb2_up = xl.load_workbook('NNSTDASU_шаблон.xlsx')
    sheet_up = wb2_up.worksheets[3]
    sheet_up[join_cl] = name_Asu
    wb2_up.save('NNSTDASU_шаблон.xlsx')






