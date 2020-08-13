#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from docx import Document
import pandas as pd
import PySimpleGUI as sg
import Check_List


Check_List.clear_list()
sg.theme('DarkGrey2')  # please make your windows colorful
layout2 = [
      [sg.Text('Выберите папку с паспортами:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FolderBrowse()],
      [sg.Text('Выберите файл с наименованиями паспортов:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FileBrowse()],
      [sg.Submit(), sg.Cancel()]]

window2 = sg.Window('Меню', layout2)
event, values = window2.read()
window2.close()
folder_path, file_path = values[0], values[1]       # get the data from the values dictionary

book = load_workbook(file_path)
ws = book.worksheets[1]
for cell in ws["A"]:
    if cell.value is None:
        cl= cell.row
        #print(cl)
        break

num=1
mylist = []
mylistASU=[]
wb = load_workbook(file_path)
sheet = wb.worksheets[1]
for val in range (1,cl):
    mylist.append(sheet.cell(row=val, column=1).value)
    mylistASU.append(sheet.cell(row=val, column=2).value)

sg.theme('DarkGrey2')
progressbar = [
    [sg.ProgressBar(len(mylist), orientation='h', size=(51, 10), key='progressbar')]
]
outputwin = [
    [sg.Output(size=(78,20))]
]

layout = [
    [sg.Frame('Progress',layout= progressbar)],
    [sg.Frame('Output', layout = outputwin)],
    [sg.Submit('Start'),sg.Cancel()]
]

window = sg.Window('Custom Progress Meter', layout)
progress_bar = window['progressbar']
while True:
    event, values = window.read(timeout=10)
    if event == 'Cancel'  or event is None:
        break
    elif event == 'Start':
        print("****** START ******")
        try:
            for d, item in enumerate(mylist):
                doc_val = mylist[d]
                name_Asu = mylistASU[d]

                first_cell_string = 'Производитель'
                name_doc= folder_path + '/' + doc_val
                document = Document(name_doc)
                # create a list of all of the table object with text of the
                # first cell equal to `first_cell_string`
                tables = [t for t in document.tables
                          if t.cell(0, 1).text.strip() == first_cell_string]

                # in the case that more than one table is found
                for table in tables:
                    data = []
                    for i, row in enumerate(table.rows):
                        text = (cell.text for cell in row.cells)
                        if i == 0:
                            keys = tuple(text)
                            continue

                        row_data = dict(zip(keys, text))
                        data.append(row_data)
                    df = pd.DataFrame(data)
                    with pd.ExcelWriter("output.xlsx", engine="openpyxl", mode="a") as writer:
                        df.to_excel(writer, startrow=0, startcol=0)
                Check_List.check_list(name_Asu)
                print(num, ") ", name_Asu, " - Done")
                num=num + 1
                progress_bar.UpdateBar(d + 1)
            print("****** FINISH ******")
        except:
            print("***** SHIT,ERROR ******")
window.close()

