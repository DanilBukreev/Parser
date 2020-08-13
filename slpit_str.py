#!/usr/bin/python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
import test4_del_sheets

def split(k):
    excel_file = 'output.xlsx'
    movies = pd.read_excel(excel_file, sheet_name=k)
    df = pd.DataFrame(movies)
    cel_nan = df.empty
    if cel_nan == True:
        df = pd.DataFrame({"Производитель": ['empty table'],
                           "Установленное ПО": ['empty table']})

        writer = pd.ExcelWriter('output2.xlsx', engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output2.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output2.xlsx')

    else:
        df.fillna('no data', inplace=True)
        new_df = pd.DataFrame(df['Установленное ПО'].str.split(',').tolist(), index=df['Производитель']).stack()
        new_df = new_df.reset_index([0, 'Производитель'])
        new_df.columns = ['Производитель', 'Установленное ПО']

        writer = pd.ExcelWriter('output2.xlsx', engine='xlsxwriter')
        new_df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output2.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output2.xlsx')


def check_null_mid(k):
    excel_file = 'output.xlsx'
    movies = pd.read_excel(excel_file, sheet_name=k)
    df = pd.DataFrame(movies)
    cel_nan = df.empty
    if cel_nan == True:
        df = pd.DataFrame({"Производитель": ['empty table'],
                           "Установленное ПО": ['empty table']})

        writer = pd.ExcelWriter('output3.xlsx', engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output3.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output3.xlsx')
    else:
        df.fillna('no data', inplace=True)
        writer = pd.ExcelWriter('output3.xlsx', engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output3.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output3.xlsx')


def check_null_set(k):
    excel_file = 'output.xlsx'
    movies = pd.read_excel(excel_file, sheet_name=k)
    df = pd.DataFrame(movies)
    cel_nan = df.empty
    if cel_nan == True:
        df = pd.DataFrame({"Производитель": ['empty table'],
                           "Установленное ПО": ['empty table']})
        writer = pd.ExcelWriter('output4.xlsx', engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output4.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output4.xlsx')
    else:
        df.fillna('no data', inplace=True)
        writer = pd.ExcelWriter('output4.xlsx', engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.bookworksheet = writer.sheets['Sheet1']
        writer.save()
        wb = load_workbook("output4.xlsx")
        ws = wb.worksheets[0]
        ws.delete_rows(1)
        wb.save('output4.xlsx')