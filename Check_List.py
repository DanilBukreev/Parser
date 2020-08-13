#!/usr/bin/python3
# -*- coding: utf-8 -*-

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
import openpyxl as xl
from openpyxl import load_workbook
import test4_del_sheets
import slpit_str



def clear_list():
    wb = load_workbook(filename='NNSTDASU_шаблон.xlsx')
    ws = wb.worksheets[1]
    thin_border = Border(top=Side(style='thin'),
                         bottom=Side(style='thin'),
                         left=Side(style='thin'),
                         right=Side(style='thin'),
                         )
    WhiteFill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')
    for row in ws['B3:E2000']:
        for cell in row:
            cell.value = None
            cell.border = thin_border
            cell.fill = WhiteFill

    wb.save('NNSTDASU_шаблон.xlsx')



def check_list(name_Asu):

    wb = load_workbook("output.xlsx")
    sheets = wb.sheetnames
    up='IP (MAC)'
    mid='Кол-во'
    set='Наименование'
    arr=[1,1,1]

    for k in range(1,len(sheets)):
        if wb[sheets[k]]['E1'].value == up:
            arr[0]=0
            ws_up=wb.worksheets[k]
            ws_up.delete_cols(1, 2)
            ws_up.delete_cols(2, 3)
            wb.save('output.xlsx')
            slpit_str.split(k)

            book = load_workbook("NNSTDASU_шаблон.xlsx")
            ws = book.worksheets[1]
            for cell in ws["C"]:
                if cell.value is None:
                    #print(cell.row )
                    break

            cl_asu = str(cell.row)
            join_asu = ''.join(['B', cl_asu])
            ws[join_asu]=name_Asu
            #cell format underline last full row
            thin_border = Border(top=Side(style='medium'))
            for i in range(1, 6):
                ws.cell(row=cell.row, column=i).border = thin_border
            book.save('NNSTDASU_шаблон.xlsx')

            # opening the source excel file
            filename = "output2.xlsx"
            wb1 = xl.load_workbook(filename)
            ws1 = wb1.worksheets[0]

            # opening the destination excel file
            filename1 = "NNSTDASU_шаблон.xlsx"
            wb2 = xl.load_workbook(filename1)
            ws2 = wb2.worksheets[1]

            # calculate total number of rows and
            # columns in source excel file
            mr = ws1.max_row
            mc = ws1.max_column

            # copying the cell values from source
            # excel file to destination excel file

            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = ws1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    ws2.cell(row=i+cell.row-1, column=j+2).value = c.value

            # saving the destination excel file
            wb2.save(str(filename1))
            test4_del_sheets.del_sheet()


        elif wb[sheets[k]]['E1'].value == mid:
            arr[1]=0
            ws_mid=wb.worksheets[k]
            ws_mid.delete_cols(1, 2)
            ws_mid.delete_cols(4, 5)
            wb.save('output.xlsx')
            slpit_str.check_null_mid(k)

            book = load_workbook("NNSTDASU_шаблон.xlsx")
            ws = book.worksheets[2]
            for cell in ws["C"]:
                if cell.value is None:
                    #print(cell.row )
                    break

            cl_asu = str(cell.row)
            join_asu = ''.join(['B', cl_asu])
            ws[join_asu]=name_Asu
            #cell format underline last full row
            thin_border = Border(top=Side(style='medium'))
            for i in range(1, 6):
                ws.cell(row=cell.row, column=i).border = thin_border
                book.save('NNSTDASU_шаблон.xlsx')

            # opening the source excel file
            filename = "output3.xlsx"
            wb1 = xl.load_workbook(filename)
            ws1 = wb1.worksheets[0]

            # opening the destination excel file
            filename1 = "NNSTDASU_шаблон.xlsx"
            wb2 = xl.load_workbook(filename1)
            ws2 = wb2.worksheets[2]

            # calculate total number of rows and
            # columns in source excel file
            mr = ws1.max_row
            mc = ws1.max_column

            # copying the cell values from source
            # excel file to destination excel file

            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = ws1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    ws2.cell(row=i+cell.row-1, column=j+2).value = c.value

            # saving the destination excel file
            wb2.save(str(filename1))
            test4_del_sheets.del_sheet()

        elif wb[sheets[k]]['E1'].value == set:
            arr[2]=0
            ws_set=wb.worksheets[k]
            ws_set.delete_cols(1, 2)
            ws_set.delete_cols(3)
            ws_set.delete_cols(4)
            wb.save('output.xlsx')
            slpit_str.check_null_set(k)

            book = load_workbook("NNSTDASU_шаблон.xlsx")
            ws = book.worksheets[3]
            for cell in ws["C"]:
                if cell.value is None:
                    #print(cell.row )
                    break

            cl_asu = str(cell.row)
            join_asu = ''.join(['B', cl_asu])
            ws[join_asu]=name_Asu

            #cell format underline last full row
            thin_border = Border(top=Side(style='medium'))
            for i in range(1, 9):
                ws.cell(row=cell.row, column=i).border = thin_border
                book.save('NNSTDASU_шаблон.xlsx')

            # opening the source excel file
            filename = "output4.xlsx"
            wb1 = xl.load_workbook(filename)
            ws1 = wb1.worksheets[0]

            # opening the destination excel file
            filename1 = "NNSTDASU_шаблон.xlsx"
            wb2 = xl.load_workbook(filename1)
            ws2 = wb2.worksheets[3]

            # calculate total number of rows and
            # columns in source excel file
            mr = ws1.max_row
            mc = ws1.max_column

            # copying the cell values from source
            # excel file to destination excel file

            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = ws1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    ws2.cell(row=i+cell.row-1, column=j+2).value = c.value

            # saving the destination excel file
            wb2.save(str(filename1))
            test4_del_sheets.del_sheet()
        else:
            print("no data! Need checking.")
            test4_del_sheets.del_sheet()

    if arr[0]==1:
        book = load_workbook("NNSTDASU_шаблон.xlsx")
        ws = book.worksheets[1]
        for cell in ws["C"]:
            if cell.value is None:
                #print(cell.row)
                break

        cl_asu = str(cell.row)
        join_asu = ''.join(['B', cl_asu])
        join_data= ''.join(['C', cl_asu])
        join_dataD = ''.join(['D', cl_asu])
        ws[join_asu] = name_Asu
        ws[join_data] = 'no table in file '
        ws[join_dataD] = 'no table in file '
        thin_border = Border(top=Side(style='medium'))
        for i in range(1, 6):
            ws.cell(row=cell.row, column=i).border = thin_border
        redFill = PatternFill(start_color='FFCC99',
                              end_color='FFCC99',
                              fill_type='solid')
        ws[join_asu].fill=redFill
        ws[join_data].fill=redFill
        ws[join_dataD].fill=redFill
        book.save('NNSTDASU_шаблон.xlsx')
    if arr[1]==1:
        book = load_workbook("NNSTDASU_шаблон.xlsx")
        ws = book.worksheets[2]
        for cell in ws["C"]:
            if cell.value is None:
                #print(cell.row)
                break

        cl_asu = str(cell.row)
        join_asu = ''.join(['B', cl_asu])
        join_data= ''.join(['C', cl_asu])
        join_dataD = ''.join(['D', cl_asu])
        ws[join_asu] = name_Asu
        ws[join_data] = 'no table in file '
        ws[join_dataD] = 'no table in file '

        thin_border = Border(top=Side(style='medium'))
        for i in range(1, 6):
            ws.cell(row=cell.row, column=i).border = thin_border
        redFill = PatternFill(start_color='FFCC99',
                              end_color='FFCC99',
                              fill_type='solid')
        ws[join_asu].fill=redFill
        ws[join_data].fill=redFill
        ws[join_dataD].fill=redFill
        book.save('NNSTDASU_шаблон.xlsx')
    if arr[2]==1:
        book = load_workbook("NNSTDASU_шаблон.xlsx")
        ws = book.worksheets[3]
        for cell in ws["C"]:
            if cell.value is None:
                #print(cell.row)
                break

        cl_asu = str(cell.row)
        join_asu = ''.join(['B', cl_asu])
        join_data= ''.join(['C', cl_asu])
        join_dataD = ''.join(['D', cl_asu])
        ws[join_asu] = name_Asu
        ws[join_data] = 'no table in file '
        ws[join_dataD] = 'no table in file '

        thin_border = Border(top=Side(style='medium'))
        for i in range(1, 9):
            ws.cell(row=cell.row, column=i).border = thin_border
        redFill = PatternFill(start_color='FFCC99',
                              end_color='FFCC99',
                              fill_type='solid')
        ws[join_asu].fill=redFill
        ws[join_data].fill=redFill
        ws[join_dataD].fill=redFill
        book.save('NNSTDASU_шаблон.xlsx')



#name_Asu='асутп'
#check_list(name_Asu)
#pyinstaller -w CountElement.pyw


